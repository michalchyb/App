from datetime import datetime
from flask import Blueprint
import openpyxl 
from typing import List, Union

shares = Blueprint('market', __name__, url_prefix='/market')

@shares.route('/shares')
def live():
    shares = read_excel_file_and_process_shares();
    return 'market_test', 200


class Share:
    def __init__(self, date: Union[datetime, str], ticker: str, money: str):
        self.date = date
        self.ticker = ticker
        self.money = money
            
    # Date property
    @property
    def date(self) -> datetime.date:
        return self._date

    @date.setter
    def date(self, value: Union[datetime, str]):
        self._date = self.convert_to_date(value)

    @property
    def ticker(self) -> str:
        return self._ticker
    
    @ticker.setter
    def ticker(self, value: str):
        self._ticker = value

    @property
    def money(self) -> float:
        return self._money

    @money.setter
    def money(self, value: str):
        self._money = self.validate_money(value)
   

    @staticmethod
    def validate_money(value: str) -> float:
        """Validate and convert the money value to a float."""
        try:
            money_value = float(value)
            if money_value <= 0:
                raise ValueError(f"Money value must be positive. Got: {money_value}")
            return money_value
        except ValueError:
            raise ValueError(f"Invalid money value: {value}. It must be a positive number.")
        
    @staticmethod
    def convert_to_date(value: Union[datetime, str]) -> datetime.date:
        """Convert string or datetime object to a date object."""
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, str):
            for date_format in ('%Y-%m-%d', '%d.%m.%Y'):
                try:
                    return datetime.strptime(value, date_format).date()
                except ValueError:
                    continue
            raise ValueError(f"Invalid date format for date: {value}. Expected format: YYYY-MM-DD or DD.MM.YYYY.")


    def __repr__(self):
        return f"Share(date={self.date}, ticker='{self.ticker}', money={self.money})"

def read_excel_file_and_process_shares() -> List[Share]:
    """Read Excel file and process share data."""
    file_path = r"C:\Users\macy\VisualStudio\app\market.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["shares"]
    return process_sheet(sheet)


def load_workbook(file_path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(file_path)

def process_sheet(sheet) -> List[Share]:
    """Process the worksheet and return a list of Share objects.""" 
    shares_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        share = create_share_from_row(row)
        shares_list.append(share)
    return shares_list

def create_share_from_row(row: tuple) -> Share:
    """Create a Share object from a row of data."""
    date_str, ticker, money = row
    return Share(date=date_str, ticker=ticker, money=money)

